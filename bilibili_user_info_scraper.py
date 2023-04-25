# 将用户ID保存到名为user_ids.txt的文件中，每行一个。
# 运行此脚本后，您将在名为output.xlsx的Excel文件中看到提取的数据。
import os
import re
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import time
import random

def get_user_data(driver, user_id):
    user_url = f"https://space.bilibili.com/{user_id}/video"
    driver.get(user_url)

    wait = WebDriverWait(driver, 3)

    def get_attribute_value1(text):
        try:
            xpath = f"//*[contains(text(), '{text}')]/parent::div"
            element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            match = re.search(r'\d+(?:,\d+)*', element.get_attribute('title'))
            return match.group()
        except TimeoutException:
            return '-1'

    def get_attribute_value(selector, attribute):
        try:
            element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
            return element.get_attribute(attribute)
        except TimeoutException:
            return '-1'

    nickname = get_attribute_value("span#h-name", "innerText")
    description = get_attribute_value("h4.h-sign", "title")
    following = get_attribute_value(".n-data.n-gz", "title")
    fans = get_attribute_value(".n-data.n-fs", "title")
    likes = get_attribute_value1(driver, "获赞数")
    views = get_attribute_value1(driver, "播放数")
    videos = get_attribute_value("li.contribution-item.cur span.num", "innerText")
    read_count = get_attribute_value1("阅读数")

    return {
        'nickname': nickname,
        'description': description,
        'following': int(str(following).replace(",", "")),
        'fans': int(str(fans).replace(",", "")),
        'likes': int(str(likes).replace(",", "")),
        'views': int(str(views).replace(",", "")),
        'read_count':int(str(read_count).replace(",", "")),
        'videos': int(str(videos).replace(",", "")),
    }

def main():
    num_count = 0
    # 从 txt 文件读取用户ID
    with open('user_ids.txt', 'r') as f:
        user_ids = [line.strip() for line in f.readlines()]

    # 检查 output.xlsx 是否存在，创建或加载工作表
    if not os.path.exists('output.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', '昵称', '个人简介', '关注数', '粉丝数', '获赞数', '播放数', '阅读数', '视频投稿数'])
    else:
        wb = load_workbook('output.xlsx')
        ws = wb.active
        num_count = ws.max_row - 1

    # 启动浏览器
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    url = "https://space.bilibili.com"
    driver.get(url)
    input("请登录，登录完成后（页面跳转）按回车键继续...")

    # 遍历用户ID并获取数据
    for user_id in user_ids[num_count:]:
        user_data = get_user_data(driver, user_id)
        row_data = [user_id] + list(user_data.values())
        ws.append(row_data)
        num_count = num_count + 1
        print(f"第{num_count}个用户已完成爬取，UID为{user_id}")

        # 保存数据到 Excel 文件
        wb.save('output.xlsx')

        # 添加随机延时
        sleep_time = random.uniform(1, 2)  # 随机生成1到2秒之间的延时
        time.sleep(sleep_time)  # 暂停执行指定的秒数

    # 关闭浏览器
    driver.quit()

    print(str("输出结果中-1代表缺省，如果表格里看到了-1，那就是很可惜，这个单元格的值没爬成功"))

if __name__ == '__main__':
    main()
